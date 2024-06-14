from openpyxl import load_workbook,Workbook
from datetime import datetime, timedelta  

 
wb = load_workbook(r'c:\\test\\Function_Check_new.xlsx',data_only=True)
  
ws = wb['From IP21 Last Data Point']
all_data=[]
for row in ws.iter_rows(min_row=3, max_col=2, values_only=True):   
    list_row = list(row)
    all_data.append([list_row[0][0:10],list_row[1]])

column1_data = []  
row_num = 10
ws2 = wb['Calculation']
flat_list=[]


data = ws2['A10:A150']
for row in data:
    for cell in row:
        if cell.value is not None:
            date_part = cell.value.split('T')[0]
            flat_list.append(date_part)


for row in flat_list:
    if row:
        try:
            date_obj = datetime.strptime(row, '%Y-%m-%d')
            new_date_obj = date_obj - timedelta(days=1)
            column1_value = new_date_obj.strftime('%Y-%m-%d')

            if not any(sublist[0] == column1_value for sublist in all_data):
                ws2.cell(row=row_num, column=7).value = ''
        except ValueError as e:
            print(e)
            pass
  
        row_num += 1  



wb.save(r'c:\\test\\Function_Check_new_tmp.xlsx')